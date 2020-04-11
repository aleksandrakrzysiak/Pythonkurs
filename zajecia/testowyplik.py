import openpyxl
import requests
import bs4

import requests
import bs4
import os

def pobieracz_obrazkow(adres):
    try:
        os.mkdir('obrazki')
    except:
        pass
    strona = requests.get(adres)
    parser = bs4.BeautifulSoup(strona.text, 'html.parser')
    obrazki = parser.find_all('img')

    for obrazek in obrazki:
       adres_obrazka = obrazek.get('src')
       nazwa_obrazka = obrazek.get('alt')
       if nazwa_obrazka is not None:
            nazwa_obrazka = nazwa_obrazka.replace(' ', '-').replace('/', '')
            with open(f'obrazki/{nazwa_obrazka}.jpg', 'wb') as jpg:
                dane_obrazka = requests.get(adres_obrazka).content
                jpg.write(dane_obrazka)
#
# site = requests.get('http://lipsum.pl/index.php?tekst=1&typ=1&dlugosc=1&slow=180&ile=1&id=5662674&html=1')
# parser = bs4.BeautifulSoup(site.text, 'html.parser')
#
# generator = parser.get('alt')
# print(generator)
# print(data.text)

# import bs4
# import requests
#
# olx_html = requests.get('https://www.olx.pl/motoryzacja/samochody/')
# parser = bs4.BeautifulSoup(olx_html.text, 'html.parser')
# obrazki = parser.find_all('img')
# for obrazek in obrazki:
#     adres = obrazek.get('alt')
#     nazwa = obrazek.get('src')
#     print(nazwa)
#     print(obrazek)
#     obrazek = requests.get('https://www.olx.pl/motoryzacja/samochody/').content
#     plik = open(f'obrazki/{nazwa}.jpg', 'wb')
#     plik.write(obrazek)
#     plik.close()

    #tworze katalog obrazki
    #sciagam obrezek z internetu
    #zapisuje jego zawrtosci do pliku NAZWA.JPG
    #gdzie NAZWA to wartość a <img 'alt'='...'>
    # Podpowiedzi:
    # 1) pliki JPG to pliki binarne

# import zajecia.moduly.menu

# excel = openpyxl.Workbook()
# nazwa_pliku = 'statystyka.xlsx'
# excel.save(nazwa_pliku)
# arkusz = excel.active
# multitool = arkusz.cell(1,1)
# ilosc_multi = zajecia.moduly.menu.menu
# for row in multitool:
#     if ilosc_multi == '':
#         ilosc_multi = 1
#     else:
#         ilosc_multi = int(ilosc_multi) + 1
# for zajecia.moduly.menu.choice in zajecia.moduly.menu.zadania:
#     ilosc_program = zajecia.moduly.menu.choice.read()

# komorkaA1 = arkusz.cell(1,1)
#
# print(komorkaA1.value)
# komorkaA2 = arkusz.cell(2,1)
# komorkaA2.value = 'komorka B2'
#
# excel.save(nazwa_pliku)

# def licznik(x):
#     excel = openpyxl.Workbook()
#     nazwa_pliku = 'licznik.xlsx'
#     excel.save(nazwa_pliku)
#     arkusz = excel.active
#     komorkaA1 = arkusz.cell(1,1)
#     komorkaA1.value = x
#     print(komorkaA1.value)
#     komorkaA2 = arkusz.cell(2,1)
#     ilosc_uruchomien = int(komorkaA2)
#     try:
#         ilosc_uruchomien = int(ilosc_uruchomien)
#         ilosc_uruchomien += 1
#     except:
#         ilosc_uruchomien = 1
#
#     print(f"=={ilosc_uruchomien}==")
#     # for row in range(1,1):
#     #     for col in range(1,2):
#     #         multitool = arkusz.cell(row, col)
#     #         if zajecia.moduly.menu.menu(zajecia):
#     #             multitool.value =+ 1
#     # for row in range (2,16):
#     #     zadania = arkusz.cell(row,col)
#     #     zadania.value = row * col
#     excel.save(nazwa_pliku)
#
# statystyka()